"""
Format Converter Module

Purpose: Convert generated educational materials to multiple output formats.
Supports: CSV → Excel, Markdown → PDF/HTML/DOCX

Based on Phase 4 of the implementation plan.
"""

import os
import logging
import re
from typing import List, Optional
from pathlib import Path

logger = logging.getLogger(__name__)


class FormatConverter:
    """
    Converts educational content to various output formats.

    Supported conversions:
        - CSV → Excel (.xlsx) with formulas and formatting
        - Markdown → HTML with MathJax and Mermaid.js support
        - Markdown → PDF using HTML intermediate
        - Markdown → DOCX using pypandoc
    """

    def __init__(self, output_dir: str = "outputs/final"):
        """
        Initialize format converter.

        Args:
            output_dir: Directory to save converted files
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

        logger.info(f"FormatConverter initialized with output_dir: {output_dir}")

    def csv_to_excel(self, csv_content: str, filename: str) -> str:
        """
        Convert CSV content to Excel file with formatting.

        Args:
            csv_content: Raw CSV string
            filename: Output filename (without extension)

        Returns:
            Path to generated Excel file
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.worksheet.table import Table, TableStyleInfo
            import csv
            from io import StringIO

            logger.info(f"Converting CSV to Excel: {filename}")

            # Parse CSV
            csv_reader = csv.reader(StringIO(csv_content.strip()))
            rows = list(csv_reader)

            if not rows:
                logger.warning("Empty CSV content")
                return ""

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Quiz Data"

            # Write data
            for row_idx, row in enumerate(rows, start=1):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value.strip())

                    # Header row formatting
                    if row_idx == 1:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                ws.column_dimensions[column_letter].width = adjusted_width

            # Add table formatting
            if len(rows) > 1:
                table_ref = f"A1:{openpyxl.utils.get_column_letter(len(rows[0]))}{len(rows)}"
                table = Table(displayName="QuizTable", ref=table_ref)
                style = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
                table.tableStyleInfo = style
                ws.add_table(table)

            # Optional: Add auto-grading formula if applicable
            # Check if headers contain "Correct_Answer_Text" and we can add a "Result" column
            if rows and "Correct_Answer_Text" in rows[0]:
                try:
                    correct_idx = rows[0].index("Correct_Answer_Text") + 1
                    # Add "Student_Answer" and "Result" columns if not present
                    if "Student_Answer" not in rows[0]:
                        student_col = len(rows[0]) + 1
                        result_col = len(rows[0]) + 2

                        ws.cell(row=1, column=student_col, value="Student_Answer")
                        ws.cell(row=1, column=result_col, value="Result")

                        # Add formula for auto-grading
                        for row_idx in range(2, len(rows) + 1):
                            correct_cell = openpyxl.utils.get_column_letter(correct_idx) + str(row_idx)
                            student_cell = openpyxl.utils.get_column_letter(student_col) + str(row_idx)
                            result_cell = ws.cell(row=row_idx, column=result_col)
                            result_cell.value = f'=IF({student_cell}={correct_cell},"Correct","Incorrect")'
                except:
                    pass  # Skip if structure doesn't match

            # Save file
            output_path = os.path.join(self.output_dir, f"{filename}.xlsx")
            wb.save(output_path)

            logger.info(f"✓ Excel file created: {output_path}")
            return output_path

        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return ""
        except Exception as e:
            logger.error(f"Error converting CSV to Excel: {e}")
            return ""

    def markdown_to_html(self, md_content: str, filename: str,
                        include_scripts: bool = True) -> str:
        """
        Convert Markdown to HTML with MathJax and Mermaid.js support.

        Args:
            md_content: Markdown content
            filename: Output filename (without extension)
            include_scripts: Include MathJax and Mermaid CDN scripts

        Returns:
            Path to generated HTML file
        """
        try:
            import markdown
            from markdown.extensions import fenced_code, tables, codehilite

            logger.info(f"Converting Markdown to HTML: {filename}")

            # Convert Markdown to HTML
            md = markdown.Markdown(extensions=[
                'fenced_code',
                'tables',
                'codehilite',
                'nl2br',
                'sane_lists'
            ])

            html_body = md.convert(md_content)

            # Wrap in HTML5 template
            html_template = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            line-height: 1.6;
            max-width: 900px;
            margin: 40px auto;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        article {{
            background: white;
            padding: 40px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        h1, h2, h3 {{
            color: #2c3e50;
            margin-top: 1.5em;
        }}
        h1 {{
            border-bottom: 3px solid #3498db;
            padding-bottom: 10px;
        }}
        h2 {{
            border-bottom: 2px solid #ecf0f1;
            padding-bottom: 8px;
        }}
        table {{
            border-collapse: collapse;
            width: 100%;
            margin: 20px 0;
        }}
        th, td {{
            border: 1px solid #ddd;
            padding: 12px;
            text-align: left;
        }}
        th {{
            background-color: #3498db;
            color: white;
            font-weight: bold;
        }}
        tr:nth-child(even) {{
            background-color: #f2f2f2;
        }}
        code {{
            background-color: #f4f4f4;
            padding: 2px 6px;
            border-radius: 3px;
            font-family: 'Courier New', monospace;
        }}
        pre {{
            background-color: #2c3e50;
            color: #ecf0f1;
            padding: 15px;
            border-radius: 5px;
            overflow-x: auto;
        }}
        pre code {{
            background: none;
            color: inherit;
        }}
        blockquote {{
            border-left: 4px solid #3498db;
            padding-left: 20px;
            margin-left: 0;
            font-style: italic;
            color: #555;
        }}
        .mermaid {{
            text-align: center;
            margin: 20px 0;
        }}
    </style>
    {scripts}
</head>
<body>
    <article>
        {content}
    </article>
</body>
</html>
"""

            # Add CDN scripts if requested
            scripts = ""
            if include_scripts:
                scripts = """
    <!-- MathJax for LaTeX rendering -->
    <script src="https://polyfill.io/v3/polyfill.min.js?features=es6"></script>
    <script id="MathJax-script" async src="https://cdn.jsdelivr.net/npm/mathjax@3/es5/tex-mml-chtml.js"></script>

    <!-- Mermaid.js for diagrams -->
    <script type="module">
        import mermaid from 'https://cdn.jsdelivr.net/npm/mermaid@10/dist/mermaid.esm.min.mjs';
        mermaid.initialize({ startOnLoad: true, theme: 'default' });
    </script>
"""

            # Generate final HTML
            final_html = html_template.format(
                title=filename.replace('_', ' ').title(),
                scripts=scripts,
                content=html_body
            )

            # Save file
            output_path = os.path.join(self.output_dir, f"{filename}.html")
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(final_html)

            logger.info(f"✓ HTML file created: {output_path}")
            return output_path

        except ImportError:
            logger.error("markdown not installed. Run: pip install markdown")
            return ""
        except Exception as e:
            logger.error(f"Error converting Markdown to HTML: {e}")
            return ""

    def markdown_to_pdf(self, md_content: str, filename: str) -> str:
        """
        Convert Markdown to PDF via HTML intermediate.

        Args:
            md_content: Markdown content
            filename: Output filename (without extension)

        Returns:
            Path to generated PDF file
        """
        try:
            # First convert to HTML
            temp_html = self.markdown_to_html(md_content, f"{filename}_temp", include_scripts=False)

            if not temp_html:
                return ""

            # Try pdfkit first
            try:
                import pdfkit

                output_path = os.path.join(self.output_dir, f"{filename}.pdf")
                pdfkit.from_file(temp_html, output_path)

                # Clean up temp HTML
                os.remove(temp_html)

                logger.info(f"✓ PDF file created: {output_path}")
                return output_path

            except ImportError:
                logger.warning("pdfkit not available, trying alternative method...")

                # Alternative: Use weasyprint if available
                try:
                    from weasyprint import HTML

                    output_path = os.path.join(self.output_dir, f"{filename}.pdf")
                    HTML(temp_html).write_pdf(output_path)

                    # Clean up temp HTML
                    os.remove(temp_html)

                    logger.info(f"✓ PDF file created: {output_path}")
                    return output_path

                except ImportError:
                    logger.error("No PDF converter available. Install pdfkit or weasyprint")
                    logger.error("For pdfkit: pip install pdfkit + install wkhtmltopdf")
                    logger.error("For weasyprint: pip install weasyprint")
                    return ""

        except Exception as e:
            logger.error(f"Error converting Markdown to PDF: {e}")
            return ""

    def markdown_to_docx(self, md_content: str, filename: str) -> str:
        """
        Convert Markdown to DOCX using pypandoc.

        Args:
            md_content: Markdown content
            filename: Output filename (without extension)

        Returns:
            Path to generated DOCX file
        """
        try:
            import pypandoc

            logger.info(f"Converting Markdown to DOCX: {filename}")

            # Save temporary markdown file
            temp_md = os.path.join(self.output_dir, f"{filename}_temp.md")
            with open(temp_md, 'w', encoding='utf-8') as f:
                f.write(md_content)

            # Convert using pypandoc
            output_path = os.path.join(self.output_dir, f"{filename}.docx")
            pypandoc.convert_file(
                temp_md,
                'docx',
                outputfile=output_path,
                extra_args=['--reference-doc=default']
            )

            # Clean up temp file
            os.remove(temp_md)

            logger.info(f"✓ DOCX file created: {output_path}")
            return output_path

        except ImportError:
            logger.error("pypandoc not installed. Run: pip install pypandoc")
            logger.error("Also ensure pandoc is installed: https://pandoc.org/installing.html")
            return ""
        except Exception as e:
            logger.error(f"Error converting Markdown to DOCX: {e}")
            return ""

    def batch_convert(self, content: str, formats: List[str], base_name: str,
                     content_type: str = 'markdown') -> dict:
        """
        Convert content to multiple formats.

        Args:
            content: Raw content (CSV or Markdown)
            formats: List of target formats ('csv', 'excel', 'pdf', 'html', 'docx')
            base_name: Base filename (without extension)
            content_type: Type of input content ('csv' or 'markdown')

        Returns:
            Dictionary mapping format to output path
        """
        logger.info(f"Batch converting {base_name} to formats: {formats}")

        results = {}

        for fmt in formats:
            fmt_lower = fmt.lower()

            if fmt_lower in ['excel', 'xlsx'] and content_type == 'csv':
                path = self.csv_to_excel(content, base_name)
                if path:
                    results['excel'] = path

            elif fmt_lower == 'html':
                if content_type == 'markdown':
                    path = self.markdown_to_html(content, base_name)
                    if path:
                        results['html'] = path

            elif fmt_lower == 'pdf':
                if content_type == 'markdown':
                    path = self.markdown_to_pdf(content, base_name)
                    if path:
                        results['pdf'] = path

            elif fmt_lower in ['docx', 'word']:
                if content_type == 'markdown':
                    path = self.markdown_to_docx(content, base_name)
                    if path:
                        results['docx'] = path

        logger.info(f"✓ Batch conversion complete: {len(results)} files created")
        return results

    def convert_notebooklm_artifacts(self, artifacts: dict, base_name: str,
                                     output_formats: List[str]) -> dict:
        """
        Convert NotebookLM artifacts to specified formats.

        Args:
            artifacts: Dict from NotebookLM with keys: quiz_csv, study_guide_md, visuals_md
            base_name: Base filename prefix
            output_formats: Target formats

        Returns:
            Dictionary of all generated files
        """
        logger.info(f"Converting NotebookLM artifacts for: {base_name}")

        all_files = {}

        # Convert quiz CSV
        if 'quiz_csv' in artifacts and artifacts['quiz_csv']:
            quiz_files = self.batch_convert(
                content=artifacts['quiz_csv'],
                formats=['excel'] if 'excel' in output_formats or 'xlsx' in output_formats else [],
                base_name=f"{base_name}_quiz",
                content_type='csv'
            )
            all_files.update(quiz_files)

        # Convert study guide
        if 'study_guide_md' in artifacts and artifacts['study_guide_md']:
            guide_formats = [f for f in output_formats if f.lower() in ['pdf', 'html', 'docx', 'word']]
            guide_files = self.batch_convert(
                content=artifacts['study_guide_md'],
                formats=guide_formats,
                base_name=f"{base_name}_study_guide",
                content_type='markdown'
            )
            all_files.update(guide_files)

        # Convert visuals/handout
        if 'visuals_md' in artifacts and artifacts['visuals_md']:
            visual_formats = [f for f in output_formats if f.lower() in ['html', 'pdf']]
            visual_files = self.batch_convert(
                content=artifacts['visuals_md'],
                formats=visual_formats,
                base_name=f"{base_name}_handout",
                content_type='markdown'
            )
            all_files.update(visual_files)

        logger.info(f"✓ NotebookLM artifact conversion complete: {len(all_files)} files")
        return all_files


# Convenience function
def convert_educational_materials(artifacts: dict, output_name: str,
                                 formats: List[str] = None,
                                 output_dir: str = "outputs/final") -> dict:
    """
    Convenience function for converting educational materials.

    Args:
        artifacts: NotebookLM artifacts dictionary
        output_name: Base output filename
        formats: List of target formats (default: ['excel', 'pdf', 'html'])
        output_dir: Output directory

    Returns:
        Dictionary of generated file paths
    """
    if formats is None:
        formats = ['excel', 'pdf', 'html']

    converter = FormatConverter(output_dir=output_dir)
    return converter.convert_notebooklm_artifacts(artifacts, output_name, formats)
